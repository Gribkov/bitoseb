import openpyxl
import os
import shutil
import datetime
import pprint
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import RED
from openpyxl import styles


#--- Вывод описания и инструкции
print('''
Данный скрипт заполняет файл себестоимости жанными из выгрузки Bi вам надо перетащить сюда 2 файла в следующем порядке:
  1) Выгрузка из Bi
  2) Файл себестоимости
Скрипт создаст копию себестотмости в той же папке откуда взяли исходную себестоимость.
''')

print('Текущий каталог: ',os.getcwd())

#--- Функция ввода путей для файлов
def myFileName(welcomeText, needCopy):
    print(welcomeText)
    fin = True
    f = 0
    while fin:
     try:
      fName = input('Введите имя файла и нажмите Enter или  введите q! для выхода:')
      if fName != 'q!':
         f = openpyxl.load_workbook(fName)
         if needCopy: #создаем копию файла и возвращаем ссылку на копию
            fNewName = 'sebestsup_'+str(datetime.datetime.today().isoformat(sep='_', timespec='hours'))+'.xlsx'
            shutil.copy(fName, fNewName)
            f = openpyxl.load_workbook(fNewName)
         fin = False
      else:
         exit()
     except FileNotFoundError:
      print('Файл не найден =(')
     except openpyxl.utils.exceptions.InvalidFileException:
      print('Файл не того формата =(')
    return (f, fNewName)
#--- Вводим путь файла из bi
#пока прикроем fileNameBi = myFileName('ВВЕДИТЕ ФАЙЛ ТРЗ ИЗ Bi', False)
fileNameBi = openpyxl.load_workbook('trzc.xlsx')

#--- Вводим файл себестоимости, создаем копию с котррой и будем работать
fileNameSeb = myFileName('ВВЕДИТЕ ФАЙЛ СЕБЕСТОИМОСТИ', True)
#fileNameSeb = openpyxl.load_workbook('seb.xlsx')


#--- Вывод предварительных данных
# Состав файла Трз
# Исполнитель -  A4 (r4, c1)
# Проект - G4 (r4, c7)
# Контракт - I4 (r4, c9)
# Трудозатраты - N4 (r4, c14)
# Неделя - D4 и E4 (r4, c4) и (r4, c5)
# Корретировка - AQ4 (r4, c43)
#


listsBi = fileNameBi.sheetnames
workListBi = fileNameBi[listsBi[0]]

listsSeb = fileNameSeb[0].sheetnames
workListSeb = fileNameSeb[0][listsSeb[0]]

#Вычислим максимальное кол-во строк в выгрузке
mxBi = 1 #сюда запишем сколько строк в выгрузке из  bi
while  str(workListBi.cell(row = mxBi, column = 1).value) != 'Общий итог':
       mxBi += 1
print('mxBi =',mxBi)

#Вычислим максимальное кол-во строк в себестоимости
mxSeb = 1 #сюда запишем сколько строк в себестоимости
while  str(workListSeb.cell(row = mxSeb, column = 1).value) != 'ENDOFTRZ':
       mxSeb += 1
print('mxSeb =',mxSeb)

#Создадим справочник контрактов со справочником исполнителей и из трз
sprTrz = {}
#заполним справочник
i = 4
for i in range(4,mxBi,1):
    #Разберем строку на составляющие
    contract = workListBi.cell(row = i, column = 9).value
    weekStart = workListBi.cell(row = i, column = 4).value
    weekEnd = workListBi.cell(row=i, column=5).value
    week = str(weekStart.day).zfill(2)+'/'+str(weekStart.month).zfill(2)+'-'+str(weekEnd.day).zfill(2)+'/'+str(weekEnd.month).zfill(2)
    ispolnitel = workListBi.cell(row = i, column = 1).value

    #Создадим структуру
    sprTrz.setdefault(week, {}) #контракты
    sprTrz[week].setdefault(contract,{}) #недели
    sprTrz[week][contract].setdefault(ispolnitel, 0.0) #исполнители
    sprTrz[week][contract][ispolnitel] += workListBi.cell(row = i, column = 14).value

for j in sprTrz:
    #найдем колонку с неделей
    for jn in range(1, 150, 1):
        if str(workListSeb.cell(row = 2, column = jn).value) == str(j):
           weekColumn = jn+1 #т.к. неделя в себестоимости объединена, то надо брать на одну дальше, что бы писать в факт
    for k in sprTrz[j]:
        for l in sprTrz[j][k]:
            w = 1
            for allRows in range(1,mxSeb,1):
                if str(k) == str(workListSeb.cell(row = allRows, column = 170).value) and \
                   str(l) == str(workListSeb.cell(row = allRows, column = 2).value) and \
                   str(workListSeb.cell(row = allRows, column = weekColumn).value) == 'None':
                   print('Неделя: '+str(j)+' Контракт: '+str(k)+' Исполнитель: '+str(l) +
                         ' ТрЗ: '+str(workListSeb.cell(row = allRows, column = weekColumn).value) +
                         ' sprTrz: '+str(sprTrz[j][k][l]))
                   workListSeb.cell(row = allRows, column = weekColumn).value = sprTrz[j][k][l]
                   workListSeb.cell(row = allRows, column = weekColumn).font = Font(bold=True, color=RED)
                   print('Неделя: ' + str(j) + ' Контракт: ' + str(k) + ' Исполнитель: ' + str(l) +
                         ' ТрЗ: ' + str(workListSeb.cell(row=allRows, column=weekColumn).value) +
                         ' sprTrz: ' + str(sprTrz[j][k][l]))

fileNameSeb[0].save(fileNameSeb[1])
fileNameSeb[0].close()
#pprint.pprint(sprTrz)
