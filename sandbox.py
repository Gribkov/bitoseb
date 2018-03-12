import openpyxl
import os
import shutil
import datetime
import pprint


print('Текущий каталог: ',os.getcwd())

#--- Функция ввода путей для файлов
def myFileName(welcomeText, needCopy):
    print(welcomeText)
    fin = True
    f = 0
    while fin:
     try:
      fName = input('Введите имя файла и нажмите Enter или  введите q! для выхода:')
      if fName != 'q!':
         f = openpyxl.load_workbook(fName)
         if needCopy: #создаем копию файла и возвращаем ссылку на копию
            fNewName = 'sebestsup_'+str(datetime.datetime.today().isoformat(sep='_',timespec='minutes'))+'.xlsx'
            shutil.copy(fName, fNewName)
            f = openpyxl.load_workbook(fNewName)
         fin = False
      else:
         exit()
     except FileNotFoundError:
      print('Файл не найден =(')
     except openpyxl.utils.exceptions.InvalidFileException:
      print('Файл не того формата =(')
    return f
#--- Вводим путь файла из bi
#пока прикроем fileNameBi = myFileName('Введите имя файла выгрузки Bi и нажмите Enter:', False)
fileNameBi = openpyxl.load_workbook('trz.xlsx')
# Состав файла Трз
# Исполнитель -  A4 (r4, c1)
# Проект - G4 (r4, c7)
# Контракт - I4 (r4, c9)
# Трудозатраты - N4 (r4, c14)
# Неделя - D4 и E4 (r4, c4) и (r4, c5)
# Корретировка - AQ4 (r4, c43)
#


listsBi = fileNameBi.sheetnames
workListBi = fileNameBi[listsBi[0]]

#Вычислим максимальное кол-во строк в выгрузке
mxBi = 1 #сюда запишем сколько строк в выгрузке из  bi
while  str(workListBi.cell(row = mxBi, column = 1).value) != 'Общий итог':
       mxBi += 1

#Создадим справочник контрактов со справочником исполнителей и из трз
sprTrz = {}
#заполним справочник
i = 4
for i in range(4,mxBi,1):
    #Разберем строку на составляющие
    contract = workListBi.cell(row = i, column = 9).value
    weekStart = workListBi.cell(row = i, column = 4).value
    weekEnd = workListBi.cell(row=i, column=5).value
    week = str(weekStart.day)+'/'+str(weekStart.month)+'-'+str(weekEnd.day)+'/'+str(weekEnd.month)
    ispolnitel = workListBi.cell(row = i, column = 1).value

    #Создадим структуру
    sprTrz.setdefault(week, {}) #контракты
    sprTrz[week].setdefault(contract,{}) #недели
    sprTrz[week][contract].setdefault(ispolnitel, 0.0) #исполнители
    sprTrz[week][contract][ispolnitel] += workListBi.cell(row = i, column = 14).value

for j in sprTrz:
    print(j)
    for k in sprTrz[j]:
        for l in sprTrz[j][k]:
            print('week')

#pprint.pprint(sprTrz)